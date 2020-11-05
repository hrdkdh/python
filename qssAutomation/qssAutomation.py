import os
import re
import openpyxl
import numpy as np
import pandas as pd
import mailing
import pygetwindow as gw
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color
from openpyxl.styles import numbers, Border, Side
from time import sleep

pd.options.display.float_format = '{:,}'.format
pd.set_option('display.max_columns', 10000)    
pd.set_option('display.max_rows',10000)

#전역변수 설정
today =""
df_result = ""
ep_id = "hrdkdh"
ep_pw = input("EP 비밀번호를 입력해 주세요 : ")
root_path = r"C:\Users\POSCOUSER\python" #실사용 시 변경할 것!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
path = root_path + "\청구서"
xlsxFile = root_path+'\8월 지도실적.xlsx' #원본파일 지정. 필요시 변경하거나 아래처럼 입력받을 것
#xlsxFile = input("파일명을 입력하여 주십시오(파일이름을 복사 붙여넣으시면 편합니다.")+'.xlsx'

mail_subject = "QSS 컨설팅 지도실적 검수요청 드립니다." #"생산기술전략실 QSS 컨설팅 지도실적 검수요청 드립니다. ('20년 9월분)"
mail_content_pre = "안녕하십니까, 포스코인재창조원 김지호 입니다.\n\n**today_for_str** 포항제철소 QSS 컨설팅 수행에 따른\n용역비 청구를 위해 수행 실적에 대한 검수를 요청드립니다.\n\n"
mail_content_post = "첨부) QSS 컨설팅 용역비 청구서\n\n수정사항 있으시면 연락부탁드립니다.\n\n감사합니다. 김지호 드림.\n\n"
maii_paste_method = "image" #메일 본문에 있는 청구서 표를 이미지가 아닌 텍스트표로 넣고 싶다면 "text"로 설정할 것

######################단가 등 회사정보 입력#####################
g_info_dic=[
    {
        "company_name":"포스코(생산전략실)",
        "pic_email" : "kyungyoub@posco.com",
        "company_cost" : 1012464,
        "family" :  "그룹장/포스코 Smart Facory 기획그룹",
        "leader" :  "김상윤",
        "damd" :  "권계찬",
        "damgr" :  "리더/포스코 Smart Facory 기획그룹"
    },
    {
        "company_name":"포스코(포항제철소)",
        "pic_email" : "ksg501@posco.com",
        "company_cost" : 1012464,
        "family" : "그룹장/포스코 인사노무그룹",
        "leader" : "김태훈",
        "damd" : "조명래",
        "damgr" : "리더/포스코 인사노무그룹"
    },
    {
        "company_name":"포스코(포)혁신Hub)",
        "pic_email" : "coma007@posco.com",
        "company_cost" : 1012464,
        "family" : "그룹장/포스코 행정섭외그룹",
        "leader" : "변재오",
        "damd" : "문홍득",
        "damgr" : "리더/포스코 행정섭외그룹"
    },
    {
        "company_name":"포스코(광양제철소)",
        "pic_email" : "jayjang@posco.com",
        "company_cost" : 1012464,
        "family" : "그룹장/포스코 인사노무그룹",
        "leader" : "남성곤",
        "damd" : "나태호",
        "damgr" : "리더/포스코 인사노무그룹"
    },
    {
        "company_name":"포스코(광)혁신Hub)",
        "pic_email" : "minamoto@posco.com",
        "company_cost" : 1012464,
        "family" : "그룹장/포스코 행정섭외그룹",
        "leader" : "이광수",
        "damd" : "차동관",
        "damgr" : "리더/포스코 행정섭외그룹"
    },
    {
        "company_name":"SNNC",
        "pic_email" : "seongchu921@snnc.co.kr",
        "company_cost" : 1030000,
        "family" : "그룹장/SNNC 기술개발그룹",
        "leader" : "유재형",
        "damd" : "강성철",
        "damgr" : "파트장/SNNC 기술개발그룹"
    },
    {
        "company_name":"한양로브틱스",
        "pic_email" : "rbxor623@hyrobot.com",
        "company_cost" : 1000000,
        "family" : "한양로보틱스",
        "leader" : "",
        "damd" : "강규택",
        "damgr" : "주임/한양로보틱스"
    },
    {
        "company_name":"포스코케미칼",
        "pic_email" : "eunj@poscochemical.com",
        "company_cost" : 1030000,
        "family" : "그룹장/포스코케미칼 인사혁신그룹",
        "leader" : "최두흔",
        "damd" : "박은지",
        "damgr" : "사원/포스코케미칼 인사혁신그룹"
    },
    {
        "company_name":"SPS(후반가공사업부)",
        "pic_email" : "sikim12@poscosps.com",
        "company_cost" : 1030000,
        "family" : "그룹장/포스코에스피에스 안전품질혁신그룹",
        "leader" : "임경우",
        "damd" : "김건우",
        "damgr" : "과장/포스코에스피에스 안전품질혁신그룹"
    },
    {
        "company_name":"포스코강판",
        "pic_email" : "po10407@poscocnc.com",
        "company_cost" : 950000,
        "family" : "그룹장/포스코강판 생산정비그룹",
        "leader" : "원유건",
        "damd" : "박명출",
        "damgr" : "대리/포스코강판 생산정비그룹"
    },
    {
        "company_name":"포스코에스피에스(TMC)",
        "pic_email" : "nocool@poscosps.com",
        "company_cost" : 1030000,
        "family" : "리더/포스코에스피에스 안전환경혁신섹션",
        "leader" : "최설해",
        "damd" : "장일우",
        "damgr" : "과장/포스코에스피에스 안전환경혁신섹션"
    },
    {
        "company_name":"포스코인터(미얀마)",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코에너지(인천)",
        "pic_email" : "kangsoo37@poscoenergy.com",
        "company_cost" : 1030000,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"부산E&E",
        "pic_email" : "pwb9060@busanene.co.kr",
        "company_cost" : 1030000,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코인터내셔널(본사)",
        "pic_email" : "hjkim12@poscointl.com",
        "company_cost" : 1030000,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코인터(우즈백)",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코에너지(LNG터미널)",
        "pic_email" : "kangsoo37@poscoenergy.com",
        "company_cost" : 1030000,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코건설(철구사업부)",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코인터내셔널(TMC)",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"ZPSS",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코인터내셔널(SRDC)",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코인터(PT.BIA)",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"인재창조원업무",
        "pic_email" : "",
        "company_cost" : 0,
        "family" : "",
        "leader" : "",
        "damd" : "",
        "damgr" : ""
    },
    {
        "company_name":"포스코에너지",
        "pic_email" : "kangsoo37@poscoenergy.com",
        "company_cost" : 1030000,
        "family" : "리더/포스코에너지 혁신지원섹션",
        "leader" : "유원상",
        "damd" : "정강수",
        "damgr" : "대리/포스코에너지 혁신지원섹션"
    }]
##############################################################

def printStepMsg(step, msg):
    print("                                                   ")
    print("===================================================")
    print("=======================STEP"+step+"=======================")
    print("===================================================")
    print(msg)
    print("                                                   ")

#진행할 준비가 되었는지 확인 (IE, Excel 등이 열려있다면 경고)
def step0CheckOptimization():
    global ep_id, ep_pw
    win = gw.getWindowsWithTitle("Excel")
    win2 = gw.getWindowsWithTitle("Internet Explorer")
    if len(win)>0:
        print("!!!오류 : 엑셀창이 열려있어 진행할 수 없습니다!!!")
        print("!!!열려있는 엑셀창을 닫은 다음 재시도해 주세요!!!")
        sys.exit()

    if len(win2)>0:
        print("!!!오류 : 인터넷 익스플로어창이 열려있어 진행할 수 없습니다!!!")
        print("!!!열려있는 인터넷 익스플로어창을 닫은 다음 재시도해 주세요!!!")
        sys.exit()

    print("체크완료")

#데이터 추출
def step1MakeQSSDataFrame():
    global today, df_result
    sheetList = []
    wb = openpyxl.load_workbook(xlsxFile)
    for i in wb.sheetnames:
        sheetList.append(i)

    # pandas를 이용하여 각 시트별 데이터 가져오기
    xlsx = pd.ExcelFile(xlsxFile)
    df_all = pd.read_excel(xlsx,header=4, sheet_name = None)

    for i in sheetList:
        df_all[i]["성명"]=i

    df = pd.concat(df_all)
    c_name=list(df.columns)

    # 데이터 정제
    col_rename=[]
    for r in c_name:
        d=re.sub('[-=+,#\?:^$@*\"※~%ㆍ!』\n\\‘|\[\]\<\>`\'…》]',"",r)
        col_rename.append(d)

    df.columns=col_rename

    df=df.fillna(0)
    df2=df.iloc[:,6:]
    df2.astype("str")
    df2=df2.replace(" ",0)
    df2=df2.replace("ㅎ",0)
    df2=df2.replace('\\',0)
    df2.astype("float")
    df3=df.iloc[:,0:6]
    df4=pd.concat([df3,df2],axis=1)
    delname=df4.loc[df4["요일"]==0].index
    df4.drop(delname, inplace=True)
    d_col=list(df4.columns).index("소계")
    df4.drop(df4.columns[d_col:],axis=1,inplace=True)

    df_result = df4

    year=df4.iloc[1,1].year
    month=df4.iloc[1,1].month
    today=str(year)+"년"+str(month)+"월"

#각 공장별 용역서 생성 및 개인별 시트 생성 & 회사정보 딕셔너리에 변수추가
def step2MakeInvoiceExcelFile():
    global today, df_result
    df = df_result

    col_name=df.columns[6:len(df.columns)]

    try:
        os.mkdir(path+"/"+today+"/")
    except:
        print(today+" 폴더가 이미 존재합니다. 덮어쓰는 중...")

    this_path=path+"\\"+today

    company_count = len(df.columns)-6
    for i in range(company_count):
        writer=pd.ExcelWriter(this_path+"\\"+today+"_"+col_name[i]+".xlsx", engine='openpyxl')
        
        summary=df.loc[:,['성명', '날짜', '요일', '활동지역(시/군)', '회사명 또는 지도부서명', '주요내용 및 이슈',col_name[i]]]
        summary1=summary.loc[summary[col_name[i]]>0]
        summary2=summary1.loc[:,["성명",col_name[i]]]
        summary3=summary2.groupby("성명").agg(["sum","count"],as_index=False)
        
        this_company = ""
        for g in g_info_dic:
            if g["company_name"] == col_name[i]:
                this_company_info = g
                break
        try:
            summary3["단가(월/일)"]=this_company_info["company_cost"]
            summary3["단가(원/시간)"]=this_company_info["company_cost"]/8
            summary3["공급가액(원)"]=summary3[col_name[i]]["count"]*summary3["단가(월/일)"]
            summary3["부가세(원)"]=summary3["공급가액(원)"]*0.1
            summary3["합계(원)"]=summary3["공급가액(원)"]+summary3["부가세(원)"]
            summary3=summary3.astype(float)
        except:
            summary3["단가(월/일)"]="0"
            summary3["단가(원/시간)"]="0"
            summary3["공급가액(원)"]="0"
            summary3["부가세(원)"]="0"
            summary3["합계(원)"]=summary3["공급가액(원)"]+summary3["부가세(원)"]
            summary3=summary3.astype(float)
        
        # 엑셀 저장
        summary3.to_excel(writer,sheet_name="용역서")
        print("엑셀파일 생성 [" + str(i+1) + "/" + str(company_count) + "] : " + col_name[i])
        
        name=summary3.index
        name_count=len(summary3.index)
        
        for n in name:
            mts=summary.loc[summary["성명"]==n,:]
            mts=mts.replace(0,"")
            mts.to_excel(writer,sheet_name=n)
            writer.save()
            writer.close()

#엑셀 수식 꾸미기
def step3DesignInvoiceExcelFile():
    #폴더 내 파일을 리스트로 변환
    this_path=path+"\\"+today
    files=os.listdir(this_path)
    files_count=len(files)
    files_c=[]

    for i in files:
        a=i.replace(today+"_","")
        files_c.append(a)

    for f in range(files_count):
        #회사정보 Dictionary에서 회사정보 로드
        this_company_name = files_c[f][0:len(files_c[f])-5]
        this_company_info = ""
        for g in g_info_dic:
            if g["company_name"] == this_company_name:
                this_company_info = g
                break

        # 용역서 시트 꾸미기
        filename=files[f]
        book = openpyxl.load_workbook(this_path+"\\"+filename)
        sheetnames=book.sheetnames
        sheet=book["용역서"]
        sheet.unmerge_cells(start_row=1,start_column=2,end_row=1,end_column=3)
        sheet.delete_rows(1,2)
        sheet.insert_rows(1,3)
        sheet.insert_cols(1)
        
        sheet['a4']="부문"
        sheet['a5']=files[f][0:len(files[f])-5]
        sheet['a5'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['b4']="이름" 
        sheet['c4']="실적(시간)"
        sheet['d4']="실적(일)"
        sheet['e4']="단가(월/일)"
        sheet['f4']="단가(원/시간)"
        sheet['g4']="공급가액(원)"
        sheet['h4']="부가세(원)"
        sheet['i4']="합계(원)"
        
        col=sheet.max_column
        row=sheet.max_row
        
        sheet['a'+str(row+1)]="합계"
        sheet['b'+str(row+1)]='=counta(b5:b'+str(row)+')&"명"'
        sheet['c'+str(row+1)]='=sum(c5:c'+str(row)+')'
        sheet['d'+str(row+1)]='=sum(d5:d'+str(row)+')'
        sheet['g'+str(row+1)]='=sum(g5:g'+str(row)+')'
        sheet['h'+str(row+1)]='=sum(h5:h'+str(row)+')'
        sheet['i'+str(row+1)]='=sum(i5:i'+str(row)+')'
        
        # 셀 정렬 ,글씨체 및 색상
        cell_total=sheet['a'+str(row+1):'i'+str(row+1)]
        for cell in range(len(cell_total[0])):
            cell_total[0][cell].font=Font(size=15,bold=True)
            cell_total[0][cell].fill=PatternFill(patternType='solid', fgColor=Color('00FF6600'))
            cell_total[0][cell].alignment = Alignment(horizontal='center', vertical='center')

        # 제목 및 글씨체
        sheet['a2']="Time & Expense Report"
        sheet['a2'].font=Font(size=9,bold=False)
        sheet['c2']="㈜포스코인재창조원 QSS 컨설팅 용역비 청구"
        sheet['c2'].font=Font(size=22,bold=True)
        sheet.merge_cells(start_row=2,start_column=3,end_row=2,end_column=8)
        sheet['i3']=today
        sheet['i3'].font=Font(size=9,bold=False)
        sheet['i3'].alignment = Alignment(horizontal='right', vertical='center')
            
        cell1=sheet['a4':'i4']
        for cell in range(len(cell1[0])):
            cell1[0][cell].font=Font(size=10,bold=True)
            cell1[0][cell].fill=PatternFill(patternType='solid', fgColor=Color('0099CCFF'))
            cell1[0][cell].alignment = Alignment(horizontal='center', vertical='center')
        
        sheet['a'+str(sheet.max_row+3)]="허만재" # 인창원 그룹리더 변경시 하단 이름 변경
        sheet['a'+str(sheet.max_row)].font=Font(size=14,bold=True)
        sheet['a'+str(sheet.max_row)].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=sheet.max_row,start_column=1,end_row=sheet.max_row,end_column=4)
        sheet['a'+str(sheet.max_row+1)]="그룹장 / ㈜포스코인재창조원 QSS지원그룹"
        sheet['a'+str(sheet.max_row)].font=Font(size=11,bold=True)
        sheet['a'+str(sheet.max_row)].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=sheet.max_row,start_column=1,end_row=sheet.max_row,end_column=4)
        

        sheet['h'+str(sheet.max_row-1)]=this_company_info["leader"]
        sheet['h'+str(sheet.max_row-1)].font=Font(size=14,bold=True)
        sheet['h'+str(sheet.max_row-1)].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=sheet.max_row-1,start_column=8,end_row=sheet.max_row-1,end_column=9)
        sheet['h'+str(sheet.max_row)]=this_company_info["family"]
        sheet['h'+str(sheet.max_row)].font=Font(size=11,bold=True)
        sheet['h'+str(sheet.max_row)].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=sheet.max_row,start_column=8,end_row=sheet.max_row,end_column=9)
        
        # 열 간격 조정
        sheet.column_dimensions['a'].width = 18
        sheet.column_dimensions['b'].width = 10
        sheet.column_dimensions['c'].width = 10
        sheet.column_dimensions['d'].width = 10
        sheet.column_dimensions['e'].width = 13
        sheet.column_dimensions['f'].width = 13
        sheet.column_dimensions['g'].width = 20
        sheet.column_dimensions['h'].width = 20
        sheet.column_dimensions['i'].width = 20
        
        #테두리 설정
        THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
        
        for rng in sheet['a4':'i'+str(row)]:
            for cell in rng:
                cell.border = THIN_BORDER # 모든테두리 설정
        
        sheet.merge_cells(start_row=5,start_column=1,end_row=4+len(sheetnames)-1,end_column=1)

        #숫자 천단위 콤마 삽입
        for cell in sheet:
            for val in cell:
                val.number_format = numbers.BUILTIN_FORMATS[3]
        
        for k in range(1,len(sheetnames)):
            sheet1=book[sheetnames[k]]
            col=sheet1.max_column
            row=sheet1.max_row
            sheet1.unmerge_cells(start_row=2,start_column=1,end_row=row,end_column=1)
            sheet1.insert_rows(1,3)
            sheet1.delete_cols(1,3)
            sheet1['d4']="부서명"
            sheet1['e4']="Task"
        
            sheet1['a3']="소속:포스코인재창조원"
            sheet1['c3']="이름:"+sheetnames[k]
            sheet1['c3'].font=Font(size=9,bold=False)
            sheet1['a1']="Monthly Time Sheet"
            sheet1['a1'].alignment = Alignment(horizontal='center', vertical='center')
            sheet1['a1'].font=Font(size=22,bold=True)
            sheet1.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6)
            
            #날짜 형식 변경
            for z in range(5,sheet1.max_row+1):
                sheet1['a'+str(z)]=str(sheet1['a'+str(z)].value.month)+"/"+str(sheet1['a'+str(z)].value.day)
            
            #소계 및 합계 추가    
            sheet1['f'+str(sheet1.max_row+1)]='=sum(f5:f'+str(sheet1.max_row)+')'
            sheet1['e'+str(sheet1.max_row)]="소계"
            sheet1.merge_cells(start_row=sheet1.max_row,start_column=1,end_row=sheet1.max_row,end_column=4)
            sheet1['f'+str(sheet1.max_row+1)]='=f'+str(sheet1.max_row)+'/8'
            sheet1['e'+str(sheet1.max_row)]="합계"
            sheet1['e'+str(sheet1.max_row)].fill=PatternFill(patternType='solid', fgColor=Color('00FFFF00'))
            sheet1['f'+str(sheet1.max_row)].fill=PatternFill(patternType='solid', fgColor=Color('00FFFF00'))
            sheet1.merge_cells(start_row=sheet1.max_row,start_column=1,end_row=sheet1.max_row,end_column=4)
            
            #컨설턴트별 시트에서 불필요한 내역 삭제
            for z in range(5,sheet1.max_row):
                if sheet1['f'+str(z)].value==None:
                    sheet1['c'+str(z)]=None
                    sheet1['d'+str(z)]=None
                    sheet1['e'+str(z)]=None
            
            # 컬럼 스타일 변경
            cell1=sheet1['a4':'e4']
            for cell in range(len(cell1[0])):
                cell1[0][cell].font=Font(size=14,bold=True)
                cell1[0][cell].fill=PatternFill(patternType='solid', fgColor=Color('00FFFF99'))
                cell1[0][cell].alignment = Alignment(horizontal='center', vertical='center')
            
            sheet1['f4'].font=Font(size=14,bold=True)
            sheet1['f4'].fill=PatternFill(patternType='solid', fgColor=Color('00CCFFFF'))
            sheet1['f4'].alignment = Alignment(horizontal='center', vertical='center')
        
            # 열 간격 조정
            sheet1.column_dimensions['a'].width = 15
            sheet1.column_dimensions['b'].width = 10
            sheet1.column_dimensions['c'].width = 20
            sheet1.column_dimensions['d'].width = 25
            sheet1.column_dimensions['e'].width = 70
            sheet1.column_dimensions['f'].width = 25
            
            col1=sheet1.max_column
            row1=sheet1.max_row
            
            THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
            
            # 셀 정렬 ,글씨체 및 색상
            
            cell_total=sheet1['a5':'f'+str(row1)]
            for r in range(row1-4):
                for c in range(col1):
                    cell_total[r][c].font=Font(size=13,bold=True)
                    cell_total[r][c].fill=PatternFill(patternType='solid',fgColor=Color('00FFFFFF'))
                    cell_total[r][c].alignment = Alignment(horizontal='center', vertical='center')
        
            for rng in sheet1['a4':'f'+str(row1)]:
                for cell in rng:
                    cell.border = THIN_BORDER # 모든테두리 설정

            sheet1['a'+str(sheet1.max_row+3)]="검수자:이성근"
            sheet1['a'+str(sheet1.max_row)].font=Font(size=12,bold=True)
            sheet1['a'+str(sheet1.max_row)].alignment = Alignment(horizontal='center', vertical='center')
            sheet1.merge_cells(start_row=sheet1.max_row,start_column=1,end_row=sheet1.max_row,end_column=4)
            sheet1['a'+str(sheet1.max_row+1)]="㈜포스코인재창조원 QSS지원그룹"
            sheet1['a'+str(sheet1.max_row)].font=Font(size=11,bold=True)
            sheet1['a'+str(sheet1.max_row)].alignment = Alignment(horizontal='center', vertical='center')
            sheet1.merge_cells(start_row=sheet1.max_row,start_column=1,end_row=sheet1.max_row,end_column=4)

            sheet1['e'+str(sheet1.max_row-1)]=this_company_info["damd"]
            sheet1['e'+str(sheet1.max_row-1)].font=Font(size=12,bold=True)
            sheet1['e'+str(sheet1.max_row-1)].alignment = Alignment(horizontal='center', vertical='center')
            sheet1.merge_cells(start_row=sheet1.max_row-1,start_column=5,end_row=sheet1.max_row-1,end_column=6)
            sheet1['e'+str(sheet1.max_row)]=this_company_info["damgr"]
            sheet1['e'+str(sheet1.max_row)].font=Font(size=11,bold=True)
            sheet1['e'+str(sheet1.max_row)].alignment = Alignment(horizontal='center', vertical='center')
            sheet1.merge_cells(start_row=sheet1.max_row,start_column=5,end_row=sheet1.max_row,end_column=6)
        
        book.save(this_path+"\\"+filename)

#각 회사별 메일 발송
def step4SendEmail():
    global path, today, mail_subject, mail_content_pre, mail_content_post, maii_paste_method, ep_id, ep_pw
    driver = mailing.initDriver()
    mailing.connectEpMail(driver, ep_id, ep_pw)
    
    this_path = path+"\\"+today
    files=os.listdir(this_path)

    today_for_str = "'" + today[2:4] + "년 " + today[5:len(today)]
    mail_content_pre = mail_content_pre.replace("**today_for_str**", today_for_str)
    for pic in g_info_dic:
        if len(pic["pic_email"])>0:
            for f in files:
                if pic["company_name"] == f.split("_")[1].split(".")[0]:
                    print(pic["company_name"]+" : 메일 작성을 시작합니다.")

                    summary_info = {}
                    person_count = 0
                    person_list = ""
                    total_days = 0
                    total_cost = 0
                    
                    #메일 제목 변경
                    this_mail_subject = pic["company_name"] + " " + mail_subject + "(" + today_for_str + "분)"
                    
                    #투입인원 수, 투입인원 이름, 지도일수, 청구금액을 파일을 열어서 확인한 후 저장함
                    book = openpyxl.load_workbook(this_path+"\\"+f, data_only=True)
                    sheetnames=book.sheetnames
                    sheet=book["용역서"]
                    start_row_no=5
                    last_row_no=0
                    for i in range(sheet.max_row):
                        if "합계" in str(sheet["a"+str(i+1)].value).strip():
                            #데이터 시작 지점은 5번열, 종료지점은 str(i)열
                            last_row_no=i
                            for j in range(4, i):
                                person_count += 1
                                person_list += str(sheet["b"+str(j+1)].value) + ", "
                                total_days += sheet["d"+str(j+1)].value
                                total_cost += sheet["i"+str(j+1)].value
                            summary_info["투입인원"] = "{:,}".format(person_count)
                            summary_info["투입인원 리스트"] = person_list[0:len(person_list)-2]
                            summary_info["지도일수"] = "{:,}".format(total_days)
                            summary_info["청구금액"] = "{:,}".format(int(total_cost))
                            break
                    book.close()
                    this_mail_content_pre = mail_content_pre + "○ 투입인원 : "+summary_info["투입인원"]+"명("+summary_info["투입인원 리스트"]+")\n○ 지도일수 : "+summary_info["지도일수"]+"일\n○ 청구금액 : "+summary_info["청구금액"]+"원\n\n"

                    mailing.openMailWindow(driver, pic["company_name"])
                    mailing.attachFiles(driver, this_path+"\\"+f, pic["company_name"])
                    mailing.writeMailContents(driver, pic["pic_email"], this_mail_subject, this_mail_content_pre, mail_content_post, pic["company_name"], this_path+"\\"+f, maii_paste_method)
                    break
        else:
            print(pic["company_name"]+" : PASS(담당자 이메일 정보가 없음)")

if __name__ == "__main__":
    printStepMsg("0", "최적화 확인중...")
    step0CheckOptimization()

    printStepMsg("1", "데이터 추출중...")
    step1MakeQSSDataFrame()

    printStepMsg("2", "회사별 엑셀파일 작성중...")
    step2MakeInvoiceExcelFile()

    printStepMsg("3", "회사별 엑셀파일 디자인중...")
    step3DesignInvoiceExcelFile()

    printStepMsg("4", "회사별 메일작성 후 임시저장중...")
    step4SendEmail()

    print("모든 작업이 끝났습니다. 수고하셨습니다!")
    print("즐거운 하루 되십시오!")
    print("                                                   ")
    print("===================================================")