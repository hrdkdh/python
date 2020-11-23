from openpyxl import load_workbook
from konlpy.tag import Kkma
from collections import Counter
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import pandas as pd

data_file_name_from_clipboard = "saved_data_from_clipboard.xlsx"

def loadTextData(file_name, sheet_index, position=None, cate="s"):
    print("\n---------------------------------------------")
    print("Step1 : 데이터를 로드합니다...")

    df = pd.read_excel(data_file_name_from_clipboard, sheet_index, )

    if position:
        df.query("직급 == '" + position + "'", inplace=True)
    if cate == "s":
        col_name = "팔로워십 강점"
    elif cate == "w":
        col_name = "팔로워십 보완점"
    
    text_list = df[col_name].values.tolist()
    text_data = ""
    for txt in text_list:
        if txt != None or txt != "":
            text_data=text_data+" "+str(txt).replace("\r", "").strip()

    print("데이터 로드완료!")
    return text_data

def makeExcelFileFromClipboard():
    df = pd.read_clipboard()
    df.to_excel(excel_writer = data_file_name_from_clipboard)

def analyseText(text_data, results_file_name="results.txt"):
    print("\n---------------------------------------------")
    print("Step2 : 단어별 형태소 및 빈도를 분석합니다... 기다려 주세요")
    kkma=Kkma()
    data_pos=kkma.pos(text_data)
    data_arr=[]

    stop_words_file = open("stop_words.txt", "r", encoding="utf-8")
    stop_words = [x.replace("\n", "").strip() for x in stop_words_file.readlines()]
    stop_words_file.close()

    print("명사만 필터링하는 중...")
    for word_pos in data_pos:
        word=word_pos[0]
        pos=word_pos[1]
        if pos=="NNG" or pos=="VV" or pos=="VA": #명사만 필터링함. 동사도 포함하려면 or pos=="VV" (VA는 형용사) 추가할 것
            data_arr.append(word)

    print("단어별 발생빈도를 정렬하고 파일에 저장하는 중...")
    counter=Counter(data_arr).most_common()
    keywords_and_frequency_for_wc = {}
    keywords_and_frequency=[]

    print("한 글자 이상 단어, 빈도수 2 이상인 것만 필터링하는 중...")
    for keyword in counter:
        word=keyword[0]
        freq=keyword[1]
        if len(word)>1 and freq>2 and word not in stop_words: #한 글자 이상 단어 + 빈도수가 2 이상 + 불용어가 아닌 것만 추출
            keywords_and_frequency_for_wc[word] = freq
            keywords_and_frequency.append({"단어" : word, "빈도" : freq})

    df = pd.DataFrame(keywords_and_frequency)
    df.to_excel(excel_writer = results_file_name)
    print("형태소 및 빈도 분석 완료!")
    return keywords_and_frequency_for_wc

def makeWordCloud(keywords_and_frequency_for_wc, results_file_name="wordcloud.png"):
    print("\n---------------------------------------------")
    print("Step3 : 워드클라우드를 생성합니다...")
    if len(keywords_and_frequency_for_wc)>0:
        font_path="NanumBarunGothicBold.ttf"
        wordcloud=WordCloud(
            font_path=font_path,
            width=800,
            height=800,
            background_color="white"
        )
        wordcloud=wordcloud.generate_from_frequencies(keywords_and_frequency_for_wc)
        array=wordcloud.to_array()

        fig=plt.figure(figsize=(10, 10))
        plt.axis("off")
        plt.imshow(array, interpolation="bilinear")

        fig.savefig(results_file_name)
        print("워드클라우드 생성완료!")
        # plt.show()
    else:
        print("데이터가 없어 워드클라우드를 생성하지 않았습니다.")

if __name__ == "__main__":
    # makeExcelFileFromClipboard()
    position_arr = ["A1", "A2", "E1", "E2", "E3", "E4", "E5", "E6", "P1", "P2", "P3", "P4", "P5", "R1", "R2"]
    ws_arr = [{"영역":"강점", "sign":"s"}, {"영역":"보완점", "sign":"w"}]
    percent = 0
    max_count = len(position_arr)*len(ws_arr)
    i = 0
    for position in position_arr:
        for ws in ws_arr:
            i += 1
            percent = str(round((i/max_count)*100)) + "%"
            print("\n=============================================")
            print("=============================================")
            print(percent + " 빈도분석 작업을 시작합니다 : " + position + ", " + ws["영역"])
            print("=============================================")
            print("=============================================")
            text_data = loadTextData(data_file_name_from_clipboard, 0, position, ws["sign"])
            keywords_and_frequency_for_wc = analyseText(text_data, "./results/" + position + "_" + ws["영역"] + "_빈도분석.xlsx")
            makeWordCloud(keywords_and_frequency_for_wc, "./results/" + position + "_" + ws["영역"] + "_wordcloud.png")

    print("모든 작업이 완료되었습니다.")
