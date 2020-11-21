#-*- coding:utf-8 -*-

#### 코드에 대한 설명을 영상으로 제작하였습니다. 아래 링크를 참고하세요.    #### 
#### https://photos.app.goo.gl/yVwtuEzuXTEvusR58            ####
#### 영상 속 코드는 아래 코드와 내용이 조금 상이하지만 80% 이상 동일합니다. ####

# 0. 파이썬 코드 수정을 위해서는 윈도우 메모장도 좋지만 코드 에디터를 사용하는 것이 낫습니다. 먼저 에디터를 설치해 주세요.
#    vs code나 sublime text를 많이 사용합니다.
#    vs code 다운로드 링크 : https://code.visualstudio.com/docs/?dv=win
#    sublime text 다운로드 링크 : https://www.sublimetext.com/
#    에디터별 확장 패키지를 설치하면 더 편리하게 코드를 수정하고 실행할 수 있지만... 여기서는 설명을 생략하겠습니다.
#    !!!아나콘다에 있는 주피터 노트북을 사용한다면 따로 설치할 필요 없습니다.

# 1. 실행을 확인한 파이썬 버전은 3.7.x 입니다. 다른 버전을 설치할 경우 실행되지 않을 수 있습니다.
#    아래 링크에서 파이썬을 다운받은 후 설치해 주세요.
#    파이썬 3.7.x 다운로드 링크 : https://www.python.org/ftp/python/3.7.9/python-3.7.9-amd64.exe
#    !!!아나콘다를 사용한다면 따로 설치할 필요 없습니다. 다만, 아나콘다의 파이썬 interpreter 버전이 3.7.x인지 확인해 주세요.
#    !!!파이썬 버전이 맞지 않다 해도 정상 실행될 수 있으니 일단 go on하고, 실행되지 않는다면 3.7.x 버전으로 다시 설치해 주세요.

# 2. 본 코드에서는 비정형 데이터인 텍스트의 형태소 분석을 위해 konlpy 패키지를 사용합니다. (konlpy가 궁금하다면 구글링하세요)
#    어쨌든... 파이썬에 konlpy 패키지를 설치하려면 jdk 1.7 이상이 설치되어 있어야 합니다. 아래 링크에서 오라클 회원가입 후 다운받아 주세요.
#    jdk 다운로드 링크 : https://www.oracle.com/kr/java/technologies/javase/javase-jdk8-downloads.html
#    스크롤을 쭈욱 내려 Windows x64 버전을 다운받아 설치하면 됩니다.
#    (만약 파이썬을 32bit로 설치했다면.. jdk도 32bit 버전으로 다운받아 설치해 주세요)

# 3. jdk 설치 후 윈도우 환경변수에서 JAVA_HOME 및 Path 설정을 꼭 하여 주세요. 별로 어렵지 않습니다.
#    윈도우 JAVA_HOME 및 Path 설정방법 안내 링크 : https://prolite.tistory.com/975
#    !!!링크에 설명된 JAVA_HOME 링크는 여러분이 설치한 자바 버전에 따라 다를 수 있습니다.
#    !!!반드시 파일탐색기로 C:\Program Files\Java 폴더에 들어간 다음 jdk-1.x.x_xxx 정보를 확인한 후 그에 맞게 JAVA_HOME 경로를 입력해 주세요.

# 4. jdk까지 설치되었다면 openpyxl, Jpype1, konlpy, collections, matplotlib, wordcloud 패키지를 설치하여야 합니다.
#    아래 명령어를 커멘드 라인에 차례로 입력한 후 설치해 주세요. (커멘드 창은 윈도우키 → cmd 입력 → 엔터하여 실행)
#    pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org openpyxl
#    pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org Jpype1
#    pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org konlpy
#    pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org collections
#    pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org matplotlib
#    pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org wordcloud
     
#    !!!잠깐! Jpype1 설치 시 아래와 같은 오류가 나온다면?
#    !!![error: Microsoft Visual C++ 14.0 is required. Get it with "Build Tools for Visual Studio"]
#    !!!이 링크를 참고하여 C++ 빌드툴을 설치하세요 : https://m.blog.naver.com/beacon71/221872094394

#    !!!아나콘다에서 설치할 경우 아나콘다 내에서 패키지를 설치하여야 합니다.
#    !!!아나콘다에서 패키지를 설치하는 방법은 아래 링크를 참고해 주세요.
#    !!!링크 : https://m.blog.naver.com/PostView.nhn?blogId=kiddwannabe&logNo=221194451967&proxyReferer=https:%2F%2Fwww.google.com%2F

# 5. 파이썬, jdk, 패키지 설치가 모두 끝났다면 본 코드 파일을 수정하여 파일명과 시트명, 데이터 영역을 설정해 주세요.
#    코드 하단의 [if __name__ == "__main__":] 블럭 내에 있는 loadTextData() 함수로 이동하여 파라미터를 여러분에 맞게 수정하면 됩니다.
#    파라미터 순서대로 파일명, 시트명, 셀영역 정보를 입력해 주세요.
#    !!!파라미터 입력 시에는 파라미터 내용을 반드시 따옴표로 감싸 주세요!
#    !!!엑셀 파일이 암호화되어 있을 경우 오류가 발생합니다. 복호화하여 사용하거나, 복호화가 어렵다면 pandas의 pd.clipboard() 함수를 이용해 데이터를 로드하는 방법을 알아보세요.

# 6. 이제 마지막입니다. 코드 파일(textAnalysis.py)이 있는 폴더로 이동해 코드를 실행해 주세요.
#    1) 커멘드 창을 열고(윈도우키 → cmd 입력 → 엔터하여 커멘드창 실행), 아래 명령어를 입력해 주세요.
#       cd 코드파일이 있는 폴더명(ex - C:\Users\POSCOUSER\Documents\textAnalysis) → 엔터
#    2) python textAnalysis.py → 엔터
#    3) 폴더에 results.txt와 wordcloud.png 파일이 생성되었는지 확인
#    !!!아나콘다에서 실행할 경우에는 주피터 노트북에서 노트북을 하나 생성한 후 코드 복붙하여 그냥 실행(run)하면 됩니다.

# 7. 참고사항 : 윈도우 환경에서는 konlpy의 Mecab 클래스는 사용할 수 없습니다.
#    Mecab 클래스를 사용하고 싶다면 리눅스나 맥에서 실행하세요.

# ※문의처 : hrdkdh

from openpyxl import load_workbook
from konlpy.tag import Kkma
from collections import Counter
from wordcloud import WordCloud
import matplotlib.pyplot as plt

def loadTextData(file_name, sheet_name, cell_range):
    print("\n---------------------------------------------")
    print("Step1 : 데이터를 로드합니다...")
    wb=load_workbook(file_name, data_only=True)
    ws=wb[sheet_name]
    cells=ws[cell_range]
    text_data=""
    for row in cells:
        for cell in row:
            if cell.value!=None:
                text_data=text_data+" "+str(cell.value)
    print("데이터 로드완료!")
    return text_data

def analyseText(text_data, results_file_name="results.txt"):
    print("\n---------------------------------------------")
    print("Step2 : 단어별 형태소 및 빈도를 분석합니다... 기다려 주세요")
    kkma=Kkma()
    data_pos=kkma.pos(text_data)
    data_arr=[]
    print("명사만 필터링하는 중...")
    for word_pos in data_pos:
        word=word_pos[0]
        pos=word_pos[1]
        if pos=="NNG" : #명사만 필터링함. 동사도 포함하려면 or pos=="VA" 추가할 것
            data_arr.append(word)

    print("단어별 발생빈도를 정렬하고 파일에 저장하는 중...")
    counter=Counter(data_arr).most_common()
    keywords_and_frequency={}
    results_file=open(results_file_name, "w", encoding="utf-8")

    print("한 글자 이상 단어, 빈도수 2 이상인 것만 필터링하는 중...")
    for keyword in counter:
        word=keyword[0]
        freq=keyword[1]
        if len(word)>1 and freq>2: #한 글자 이상 단어 + 빈도수가 2 이상인 것만 추출
            keywords_and_frequency[word]=freq
            this_text=word+" : "+str(freq)+"건\n"
            results_file.write(this_text)

    results_file.close()
    print("형태소 및 빈도 분석 완료!")
    return keywords_and_frequency

def makeWordCloud(keywords_and_frequency, results_file_name="wordcloud.png"):
    print("\n---------------------------------------------")
    print("Step3 : 워드클라우드를 생성합니다...")
    font_path="NanumBarunGothicBold.ttf"
    wordcloud=WordCloud(
        font_path=font_path,
        width=800,
        height=800,
        background_color="white"
    )
    wordcloud=wordcloud.generate_from_frequencies(keywords_and_frequency)
    array=wordcloud.to_array()

    fig=plt.figure(figsize=(10, 10))
    plt.axis("off")
    plt.imshow(array, interpolation="bilinear")

    fig.savefig(results_file_name)
    print("워드클라우드 생성완료!")
    plt.show()

if __name__ == "__main__":
    #Step1 : 엑셀파일에 있는 텍스트를 불러와 한줄짜리 String으로 붙여 text_data 변수(String)로 출력해 준다.
    #파라미터 설명 : 분석할 엑셀파일명, 엑셀파일에 있는 시트명, 텍스트가 있는 셀의 범위
    text_data = loadTextData("sample_data.xlsx", "VOC결과", "C2:C415")

    #Step2 : text_data에 있는 단어를 형태소별로 분리하고, 각 단어별 빈도수를 계산한 결과를 txt파일로 저장 및 keywords_and_frequency 변수(딕셔너리)로 출력해 준다.
    #파라미터 설명 : 분석할 텍스트 원본, 결과를 저장할 파일명
    keywords_and_frequency = analyseText(text_data, "results.txt")

    #Step3 : keywords_and_frequency를 바탕으로 워드클라우드를 생성한 후 화면에 띄워주고 파일로 저장한다.
    #파라미터 설명 : 키워드별 빈도수가 저장된 딕셔너리, 워드클라우드 그림을 저장할 파일명
    makeWordCloud(keywords_and_frequency, "wordcloud.png")

    print("\n모든 작업이 완료되었습니다.")
	